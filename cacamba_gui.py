import tkinter as tk
from tkinter import ttk, messagebox, simpledialog, Scrollbar
import datetime
import time
import os
import sys
import json
from openpyxl import load_workbook, Workbook
from colorama import Fore, init
import requests
import folium
from urllib.parse import urlencode
from dataclasses import dataclass
from typing import List, Dict, Tuple, Optional, Any


# Inicializa o colorama para resetar cores automaticamente
init(autoreset=True)


@dataclass
class Cacamba:
    """Classe para representar uma caçamba."""
    numero: str
    cep: str
    adnumero: str
    data_colocacao: str
    rua: str
    bairro: str
    cidade: str
    uf: str
    latitude: float = None
    longitude: float = None

    @property
    def dias_no_local(self) -> int:
        """Calcula quantos dias a caçamba está no local."""
        try:
            data = datetime.datetime.strptime(self.data_colocacao, '%d/%m/%Y')
            return (datetime.datetime.now() - data).days
        except ValueError:
            return 0

    @property
    def endereco_completo(self) -> str:
        """Retorna o endereço completo formatado."""
        return f"{self.rua}, {self.adnumero}, {self.bairro}, {self.cidade}, {self.uf}"

    @property
    def precisa_retirada(self) -> bool:
        """Verifica se a caçamba precisa ser retirada (mais de 3 dias no local)."""
        return self.dias_no_local >= 3


class GerenciadorArquivos:
    """Classe para gerenciar operações de arquivo."""
    
    ARQUIVO_PADRAO = 'cacambas.xlsx'
    ARQUIVO_CONFIG = 'config.json'
    
    @staticmethod
    def obter_caminho_arquivo() -> str:
        """Retorna o caminho completo do arquivo de dados."""
        try:
            # Obtém o diretório do executável em vez do __file__ 
            # para compatibilidade com auto-py-to-exe
            if getattr(sys, 'frozen', False):
                # Executando como executável compilado
                diretorio_base = os.path.dirname(sys.executable)
            else:
                # Executando como script Python
                diretorio_base = os.path.dirname(os.path.abspath(__file__))
            
            # Verifica se existe arquivo de configuração
            caminho_config = os.path.join(diretorio_base, GerenciadorArquivos.ARQUIVO_CONFIG)
            print(Fore.CYAN + f"Procurando configuração em: {caminho_config}")
            
            if os.path.exists(caminho_config):
                with open(caminho_config, 'r') as f:
                    config = json.load(f)
                    caminho_arquivo = config.get('caminho_arquivo')
                    
                    # Verificar se o caminho existe
                    if caminho_arquivo and os.path.exists(os.path.dirname(caminho_arquivo)):
                        print(Fore.GREEN + f"Usando arquivo de dados em: {caminho_arquivo}")
                        return caminho_arquivo
                    else:
                        print(Fore.YELLOW + f"Caminho configurado não existe mais: {caminho_arquivo}")
                        # Se o diretório não existir mais, podemos perguntar novamente
                        return GerenciadorArquivos.solicitar_novo_caminho(diretorio_base)
        except Exception as e:
            print(Fore.RED + f"Erro ao ler configuração: {e}")
            
        # Caminho padrão se não houver configuração
        if getattr(sys, 'frozen', False):
            diretorio_base = os.path.dirname(sys.executable)
        else:
            diretorio_base = os.path.dirname(os.path.abspath(__file__))
            
        caminho_padrao = os.path.join(diretorio_base, GerenciadorArquivos.ARQUIVO_PADRAO)
        print(Fore.YELLOW + f"Usando caminho padrão: {caminho_padrao}")
        return caminho_padrao
    
    @staticmethod
    def solicitar_novo_caminho(diretorio_base=None) -> str:
        """Solicita um novo caminho para o arquivo caso o anterior não exista mais."""
        try:
            from tkinter import filedialog, Tk
            
            root_temp = Tk()
            root_temp.withdraw()
            
            messagebox.showinfo(
                "Localização do Arquivo",
                "O local anterior do arquivo não está mais disponível.\n"
                "Por favor, escolha um novo local para o arquivo de dados."
            )
            
            # Define o diretório inicial como o diretório do executável se não for especificado
            if diretorio_base is None:
                if getattr(sys, 'frozen', False):
                    diretorio_base = os.path.dirname(sys.executable)
                else:
                    diretorio_base = os.path.dirname(os.path.abspath(__file__))
                    
            caminho_arquivo = filedialog.asksaveasfilename(
                title="Salvar arquivo de dados das caçambas",
                defaultextension=".xlsx",
                filetypes=[("Arquivo Excel", "*.xlsx")],
                initialfile=GerenciadorArquivos.ARQUIVO_PADRAO,
                initialdir=diretorio_base
            )
            
            root_temp.destroy()
            
            if not caminho_arquivo:
                caminho_arquivo = os.path.join(diretorio_base, GerenciadorArquivos.ARQUIVO_PADRAO)
                print(Fore.YELLOW + "Usuário cancelou. Usando caminho padrão.")
            
            # Salva a nova configuração
            GerenciadorArquivos.salvar_configuracao(caminho_arquivo, diretorio_base)
            return caminho_arquivo
            
        except Exception as e:
            print(Fore.RED + f"Erro ao solicitar novo caminho: {e}")
            if diretorio_base is None:
                if getattr(sys, 'frozen', False):
                    diretorio_base = os.path.dirname(sys.executable)
                else:
                    diretorio_base = os.path.dirname(os.path.abspath(__file__))
            return os.path.join(diretorio_base, GerenciadorArquivos.ARQUIVO_PADRAO)
    
    @staticmethod
    def salvar_configuracao(caminho: str, diretorio_base=None) -> None:
        """Salva o caminho do arquivo na configuração."""
        try:
            config = {'caminho_arquivo': caminho}
            
            # Define o diretório de base para salvar a configuração
            if diretorio_base is None:
                if getattr(sys, 'frozen', False):
                    diretorio_base = os.path.dirname(sys.executable)
                else:
                    diretorio_base = os.path.dirname(os.path.abspath(__file__))
                    
            caminho_config = os.path.join(diretorio_base, GerenciadorArquivos.ARQUIVO_CONFIG)
            print(Fore.CYAN + f"Salvando configuração em: {caminho_config}")
            
            with open(caminho_config, 'w') as f:
                json.dump(config, f)
            print(Fore.GREEN + f"Configuração salva em {caminho_config}")
        except Exception as e:
            print(Fore.RED + f"Erro ao salvar configuração: {e}")
    
    @staticmethod
    def criar_arquivo_se_nao_existir() -> None:
        """Cria o arquivo Excel se não existir."""
        # Determina o diretório base - diferente para script vs executable
        if getattr(sys, 'frozen', False):
            diretorio_base = os.path.dirname(sys.executable)
        else:
            diretorio_base = os.path.dirname(os.path.abspath(__file__))
        
        caminho_config = os.path.join(diretorio_base, GerenciadorArquivos.ARQUIVO_CONFIG)
        print(Fore.CYAN + f"Verificando existência de configuração em: {caminho_config}")
        
        # Verifica se é a primeira execução
        primeira_execucao = not os.path.exists(caminho_config)
        
        if primeira_execucao:
            print(Fore.CYAN + "Primeira execução detectada. Solicitando local para salvar arquivo de dados.")
            try:
                # Solicita ao usuário o caminho para salvar o arquivo
                from tkinter import filedialog, Tk
                
                root_temp = Tk()
                root_temp.withdraw()  # Esconde a janela principal
                
                messagebox.showinfo(
                    "Configuração Inicial",
                    "Bem-vindo ao Gerenciador de Caçambas!\n\n"
                    "Este parece ser o primeiro uso do sistema.\n"
                    "Por favor, escolha onde salvar o arquivo de dados das caçambas."
                )
                
                caminho_arquivo = filedialog.asksaveasfilename(
                    title="Salvar arquivo de dados das caçambas",
                    defaultextension=".xlsx",
                    filetypes=[("Arquivo Excel", "*.xlsx")],
                    initialfile=GerenciadorArquivos.ARQUIVO_PADRAO,
                    initialdir=diretorio_base
                )
                
                root_temp.destroy()
                
                if not caminho_arquivo:
                    print(Fore.YELLOW + "Usuário cancelou a seleção de caminho. Usando caminho padrão.")
                    caminho_arquivo = os.path.join(diretorio_base, GerenciadorArquivos.ARQUIVO_PADRAO)
                
                # Salva a configuração
                GerenciadorArquivos.salvar_configuracao(caminho_arquivo, diretorio_base)
                
            except Exception as e:
                print(Fore.RED + f"Erro ao solicitar caminho do arquivo: {e}")
                caminho_arquivo = os.path.join(diretorio_base, GerenciadorArquivos.ARQUIVO_PADRAO)
        else:
            caminho_arquivo = GerenciadorArquivos.obter_caminho_arquivo()
        
        # Cria o arquivo se não existir
        if not os.path.exists(caminho_arquivo):
            try:
                wb = Workbook()
                ws = wb.active
                ws.append(['Numero', 'CEP', 'adnumero', 'data_colocacao', 'Rua', 
                          'Bairro', 'Cidade', 'UF', 'latitude', 'longitude'])
                wb.save(caminho_arquivo)
                print(Fore.GREEN + f"Arquivo {caminho_arquivo} criado com sucesso!")
            except Exception as e:
                print(Fore.RED + f"Erro ao criar o arquivo {caminho_arquivo}: {e}")
                
                # Se houver erro, tenta criar no local padrão
                caminho_padrao = os.path.join(diretorio_base, GerenciadorArquivos.ARQUIVO_PADRAO)
                try:
                    wb = Workbook()
                    ws = wb.active
                    ws.append(['Numero', 'CEP', 'adnumero', 'data_colocacao', 'Rua', 
                              'Bairro', 'Cidade', 'UF', 'latitude', 'longitude'])
                    wb.save(caminho_padrao)
                    print(Fore.YELLOW + f"Arquivo criado no local padrão: {caminho_padrao}")
                    GerenciadorArquivos.salvar_configuracao(caminho_padrao, diretorio_base)
                except Exception as e2:
                    print(Fore.RED + f"Erro ao criar o arquivo no local padrão: {e2}")
    
    @staticmethod
    def carregar_cacambas() -> List[Cacamba]:
        """Carrega os dados das caçambas do arquivo Excel."""
        GerenciadorArquivos.criar_arquivo_se_nao_existir()
        try:
            wb = load_workbook(GerenciadorArquivos.obter_caminho_arquivo())
            ws = wb.active
            cacambas = []
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0]:  # Verifica se o número da caçamba existe
                    cacamba = Cacamba(
                        numero=str(row[0]),
                        cep=str(row[1]),
                        adnumero=str(row[2]),
                        data_colocacao=str(row[3]),
                        rua=str(row[4]),
                        bairro=str(row[5]),
                        cidade=str(row[6]),
                        uf=str(row[7]),
                        latitude=row[8],
                        longitude=row[9]
                    )
                    cacambas.append(cacamba)
            
            return cacambas
        except Exception as e:
            print(Fore.RED + f"Erro ao carregar dados: {e}")
            return []
    
    @staticmethod
    def salvar_cacamba(cacamba: Cacamba) -> bool:
        """Salva uma nova caçamba no arquivo."""
        try:
            wb = load_workbook(GerenciadorArquivos.obter_caminho_arquivo())
            ws = wb.active
            
            # Verifica se o número já existe
            numeros_existentes = [str(row[0]) for row in ws.iter_rows(min_row=2, values_only=True)]
            if cacamba.numero in numeros_existentes:
                return False
            
            ws.append([
                cacamba.numero, 
                cacamba.cep, 
                cacamba.adnumero, 
                cacamba.data_colocacao, 
                cacamba.rua, 
                cacamba.bairro, 
                cacamba.cidade, 
                cacamba.uf, 
                cacamba.latitude, 
                cacamba.longitude
            ])
            
            wb.save(GerenciadorArquivos.obter_caminho_arquivo())
            return True
        except Exception as e:
            print(Fore.RED + f"Erro ao salvar caçamba: {e}")
            return False
    
    @staticmethod
    def remover_cacamba(numero: str) -> bool:
        """Remove uma caçamba do arquivo pelo número."""
        try:
            wb = load_workbook(GerenciadorArquivos.obter_caminho_arquivo())
            ws = wb.active
    
            linha_para_remover = None
            for row in ws.iter_rows(min_row=2):
                if str(row[0].value) == numero:
                    linha_para_remover = row[0].row
                    break
            
            if linha_para_remover:
                ws.delete_rows(linha_para_remover, 1)
                wb.save(GerenciadorArquivos.obter_caminho_arquivo())
                return True
            
            return False
        except Exception as e:
            print(Fore.RED + f"Erro ao remover caçamba: {e}")
            return False


class ServicoLocalizacao:
    """Classe para serviços de localização e geolocalização."""
    
    @staticmethod
    def obter_endereco_por_cep(cep: str) -> Optional[Dict[str, str]]:
        """Obtém informações de endereço a partir do CEP usando a API ViaCEP."""
        try:
            url = f"https://viacep.com.br/ws/{cep}/json/"
            resposta = requests.get(url).json()
            
            if "erro" not in resposta:
                endereco = {
                    'rua': resposta["logradouro"],
                    'bairro': resposta["bairro"],
                    'cidade': resposta["localidade"],
                    'uf': resposta["uf"]
                }
                time.sleep(1)  # Evita sobrecarga na API
                return endereco
            return None
        except Exception as e:
            print(Fore.RED + f"Erro ao obter endereço pelo CEP: {e}")
            return None

    @staticmethod
    def obter_coordenadas(endereco: str) -> Optional[Tuple[float, float]]:
        """Obtém as coordenadas geográficas a partir do endereço usando a API Nominatim."""
        try:
            params = urlencode({"q": endereco, "format": "json"})
            url = f"https://nominatim.openstreetmap.org/search?{params}"
            time.sleep(1)  # Respeita limites da API
    
            headers = {
                "User-Agent": "Mozilla/5.0 (compatible; CacambaGerenciador/1.0)"
            }
    
            resposta = requests.get(url, headers=headers).json()
    
            if resposta:
                latitude = float(resposta[0]["lat"])
                longitude = float(resposta[0]["lon"])
                return latitude, longitude
            return None
        except Exception as e:
            print(Fore.RED + f"Erro ao obter coordenadas: {e}")
            return None
    
    @staticmethod
    def gerar_mapa(cacambas: List[Cacamba]) -> None:
        """Gera um mapa interativo com as localizações das caçambas."""
        try:
            # Localização inicial do mapa (Brasil)
            mapa = folium.Map(location=[-22.9068, -43.1729], zoom_start=12)
            
            for cacamba in cacambas:
                if cacamba.latitude and cacamba.longitude:
                    # Cor do marcador baseada no tempo no local
                    cor = 'red' if cacamba.precisa_retirada else 'blue'
                    
                    # Texto com informações da caçamba
                    popup_text = f"""
                        <b>Caçamba {cacamba.numero}</b><br>
                        Endereço: {cacamba.endereco_completo}<br>
                        Data de colocação: {cacamba.data_colocacao}<br>
                        Dias no local: {cacamba.dias_no_local}
                    """
                    
                    folium.Marker(
                        location=[cacamba.latitude, cacamba.longitude],
                        popup=folium.Popup(popup_text, max_width=300),
                        icon=folium.Icon(color=cor, icon='trash')
                    ).add_to(mapa)
            
            # Obter o diretório do arquivo de dados para salvar o mapa no mesmo local
            diretorio_dados = os.path.dirname(GerenciadorArquivos.obter_caminho_arquivo())
            arquivo_mapa = os.path.join(diretorio_dados, 'mapa_cacambas.html')
            
            mapa.save(arquivo_mapa)
            print(Fore.GREEN + f"Mapa interativo salvo como '{arquivo_mapa}'")
            return arquivo_mapa
        except Exception as e:
            print(Fore.RED + f"Erro ao gerar mapa: {e}")
            return None


class ProcessadorDatas:
    """Classe para processar formatos de data."""
    
    @staticmethod
    def validar_e_formatar_data(data_texto: str) -> Optional[str]:
        """Valida e formata uma data em diversos formatos."""
        try:
            # Verifica caracteres permitidos
            if not all(c.isdigit() or c == '/' for c in data_texto):
                return None
                
            data_dt = None
            # Verifica formato com barras
            if '/' in data_texto and data_texto.count('/') == 2:
                partes = data_texto.split('/')
                if len(partes[2]) == 2:  # dd/mm/yy
                    data_dt = datetime.datetime.strptime(data_texto, '%d/%m/%y')
                else:  # dd/mm/yyyy
                    data_dt = datetime.datetime.strptime(data_texto, '%d/%m/%Y')
            # Formato sem barras
            elif data_texto.isdigit():
                if len(data_texto) == 6:  # ddmmyy
                    data_dt = datetime.datetime.strptime(data_texto, '%d%m%y')
                elif len(data_texto) == 8:  # ddmmyyyy
                    data_dt = datetime.datetime.strptime(data_texto, '%d%m%Y')
                else:
                    return None
            else:
                return None
                
            # Formata para padrão dd/mm/yyyy
            return data_dt.strftime('%d/%m/%Y')
        except ValueError:
            return None


class GerenciadorCacambas:
    """Classe principal para gerenciar caçambas."""
    
    def __init__(self):
        """Inicializa o gerenciador de caçambas."""
        self.interface = None  # Será definido posteriormente
    
    def registrar_cacamba(self, root) -> None:
        """Registra uma nova caçamba com interface gráfica."""
        # Solicita número da caçamba
        numero = simpledialog.askstring("Registrar Caçamba", 
                                       "Digite o número da caçamba:", 
                                       parent=root)
        if not numero:
            return
            
        # Verifica se já existe
        cacambas = GerenciadorArquivos.carregar_cacambas()
        numeros_existentes = [c.numero for c in cacambas]
        if numero in numeros_existentes:
            messagebox.showwarning("ALERTA", f"A caçamba {numero} já está registrada.")
            return

        # Solicita CEP
        cep = simpledialog.askstring("Registrar Caçamba", 
                                    "Digite o CEP:", 
                                    parent=root)
        if not cep:
            return
            
        # Obtém informações de endereço pelo CEP
        endereco_info = ServicoLocalizacao.obter_endereco_por_cep(cep)
        if not endereco_info:
            messagebox.showerror("Erro", "CEP inválido ou não encontrado.")
            return
            
        # Solicita número do endereço
        adnumero = simpledialog.askstring("Registrar Caçamba", 
                                         "Digite o número do endereço:", 
                                         parent=root)
        if not adnumero:
            return
            
        # Solicita data de colocação
        data_texto = simpledialog.askstring("Registrar Caçamba", 
                                          "Digite a data de colocação (dd/mm/aa):", 
                                          parent=root)
        if not data_texto:
            return
            
        # Valida e formata a data
        data_formatada = ProcessadorDatas.validar_e_formatar_data(data_texto)
        if not data_formatada:
            messagebox.showerror("Erro", 
                             "Data inválida. Formatos aceitos: dd/mm/aa, dd/mm/aaaa, ddmmaa, ddmmaaaa")
            return
            
        # Monta o endereço completo para geolocalização
        endereco_completo = (
            f"{endereco_info['rua']}, {adnumero}, "
            f"{endereco_info['bairro']}, {endereco_info['cidade']}, "
            f"{endereco_info['uf']}, Brasil"
        )
        
        # Obtém coordenadas
        coordenadas = ServicoLocalizacao.obter_coordenadas(endereco_completo)
        if not coordenadas:
            messagebox.showerror("Erro", "Não foi possível obter as coordenadas do endereço.")
            return
            
        # Cria objeto Cacamba
        nova_cacamba = Cacamba(
            numero=numero,
            cep=cep,
            adnumero=adnumero,
            data_colocacao=data_formatada,
            rua=endereco_info['rua'],
            bairro=endereco_info['bairro'],
            cidade=endereco_info['cidade'],
            uf=endereco_info['uf'],
            latitude=coordenadas[0],
            longitude=coordenadas[1]
        )
        
        # Salva a caçamba
        if GerenciadorArquivos.salvar_cacamba(nova_cacamba):
            messagebox.showinfo("Sucesso", f"Caçamba {numero} registrada com sucesso!")
            # Atualiza a interface
            if self.interface:
                self.interface.atualizar_lista_cacambas()
                self.interface.gerar_e_mostrar_mapa()
        else:
            messagebox.showerror("Erro", "Não foi possível registrar a caçamba.")
    
    def remover_cacamba(self, root) -> None:
        """Remove uma caçamba existente."""
        numero = simpledialog.askstring("Remover Caçamba", 
                                      "Digite o número da caçamba:", 
                                      parent=root)
        if not numero:
            return
            
        if GerenciadorArquivos.remover_cacamba(numero):
            messagebox.showinfo("Sucesso", f"Caçamba {numero} removida com sucesso!")
            # Atualiza a interface
            if self.interface:
                self.interface.atualizar_lista_cacambas()
                self.interface.gerar_e_mostrar_mapa()
        else:
            messagebox.showwarning("ALERTA", f"A caçamba {numero} não está registrada ou não pôde ser removida.")
    
    def verificar_cacambas_para_retirada(self) -> List[Cacamba]:
        """Verifica quais caçambas estão prontas para retirada."""
        cacambas = GerenciadorArquivos.carregar_cacambas()
        return [c for c in cacambas if c.precisa_retirada]
    
    def set_interface(self, interface) -> None:
        """Define a interface gráfica associada ao gerenciador."""
        self.interface = interface


class InterfaceGrafica:
    """Classe para gerenciar a interface gráfica."""
    
    def __init__(self, gerenciador: GerenciadorCacambas):
        """Inicializa a interface gráfica."""
        self.gerenciador = gerenciador
        self.gerenciador.set_interface(self)
        
        # Cores e estilos - atualizado com cores Chromium
        self.cores = {
            'azul': "#4285F4",  # Azul Google/Chrome
            'azul_hover': "#5C9CFF",
            'azul_pressed': "#3B78E7",
            'verde': "#34A853",  # Verde Google
            'vermelho': "#EA4335",  # Vermelho Google
            'amarelo': "#FBBC05",  # Amarelo Google
            'bg': "#F5F5F5",  # Fundo cinza claro
            'texto': "#3C4043",  # Texto escuro Google
            'borda': "#DADCE0"  # Borda Google
        }
        
        # Criação da janela principal
        self.root = tk.Tk()
        self.root.title("Gerenciador de Caçambas")
        self.root.geometry("800x600")
        self.root.configure(bg=self.cores['bg'])
        
        # Configuração de estilo
        self.configurar_estilo()
        
        # Criação da UI
        self.criar_interface()
        
        # Atualiza a lista de caçambas
        self.atualizar_lista_cacambas()
    
    def configurar_estilo(self) -> None:
        """Configura o estilo visual da aplicação no estilo Chromium."""
        style = ttk.Style()
        style.theme_use('clam')
    
        # Estilo para botões - estilo Chrome com cantos arredondados
        style.configure(
            'Chrome.TButton', 
            font=('Roboto', 10),
            padding=10,
            relief='flat',
            background=self.cores['azul'],
            foreground='white',
            borderwidth=0,
            focusthickness=0
        )
        
        style.map(
            'Chrome.TButton',
            background=[('active', self.cores['azul_hover']), 
                        ('pressed', self.cores['azul_pressed'])],
            relief=[('pressed', 'flat')]
        )
        
        # Estilo para botão vermelho
        style.configure(
            'Red.TButton', 
            font=('Roboto', 10),
            padding=10,
            relief='flat',
            background=self.cores['vermelho'],
            foreground='white',
            borderwidth=0
        )
        
        style.map(
            'Red.TButton',
            background=[('active', '#FF5A52'), 
                        ('pressed', '#D03C31')],
            relief=[('pressed', 'flat')]
        )
        
        # Estilo para frames - cantos arredondados
        style.configure(
            'Rounded.TFrame', 
            background='white',
            borderwidth=1, 
            relief='solid'
        )
                    
        # Estilo para label frames - cantos arredondados
        style.configure(
            'RoundedLabel.TLabelframe',
            background='white',
            borderwidth=1,
            relief='solid'
        )
        
        # Estilo para o título do label frame
        style.configure(
            'RoundedLabel.TLabelframe.Label', 
            font=('Roboto', 11),
            foreground=self.cores['azul'],
            background='white',
            anchor='center'
        )
        
        # Estilo para o rodapé
        style.configure(
            'Footer.TFrame', 
            background=self.cores['bg'],
            borderwidth=0,
            relief='flat'
        )
    
    def criar_interface(self) -> None:
        """Cria a interface gráfica principal com estilo Chromium."""
        # Frame principal com bordas arredondadas
        self.frame_principal = ttk.Frame(self.root, style='Rounded.TFrame', padding="15")
        self.frame_principal.pack(fill=tk.BOTH, expand=True, padx=15, pady=15)
        
        # Aplicar bordas arredondadas ao frame principal
        self._aplicar_cantos_arredondados(self.frame_principal, 10)
        
        # Título com estilo Google
        titulo = ttk.Label(
            self.frame_principal, 
            text="Gerenciador de Caçambas", 
            font=('Roboto', 18, 'bold'),
            foreground=self.cores['azul'],
            background='white'
        )
        titulo.pack(pady=(5, 20))
    
        # Frame para botões com visual chromium
        frame_botoes = ttk.Frame(self.frame_principal, style='Rounded.TFrame')
        frame_botoes.pack(fill=tk.X, pady=10)
    
        # Container para centralizar os botões
        container_botoes = ttk.Frame(frame_botoes, style='Rounded.TFrame')
        container_botoes.pack(anchor=tk.CENTER)
    
        # Botões com estilo Chromium
        self.btn_registrar = ttk.Button(
            container_botoes, 
            text="Registrar Caçamba", 
            command=lambda: self.gerenciador.registrar_cacamba(self.root),
            style='Chrome.TButton'
        )
        self.btn_remover = ttk.Button(
            container_botoes, 
            text="Remover Caçamba", 
            command=lambda: self.gerenciador.remover_cacamba(self.root),
            style='Red.TButton'
        )
        self.btn_mapa = ttk.Button(
            container_botoes, 
            text="Visualizar Mapa", 
            command=self.gerar_e_mostrar_mapa,
            style='Chrome.TButton'
        )
        
        # Adicionar bordas arredondadas aos botões
        self._aplicar_cantos_arredondados(self.btn_registrar, 20)
        self._aplicar_cantos_arredondados(self.btn_remover, 20)
        self._aplicar_cantos_arredondados(self.btn_mapa, 20)
        
        self.btn_registrar.pack(side=tk.LEFT, padx=8)
        self.btn_remover.pack(side=tk.LEFT, padx=8)
        self.btn_mapa.pack(side=tk.LEFT, padx=8)
        
        # Frame para listbox com cantos arredondados
        self.frame_lista = ttk.LabelFrame(
            self.frame_principal, 
            text="Caçambas registradas", 
            padding=12,
            style='RoundedLabel.TLabelframe'
        )
        self.frame_lista.pack(fill=tk.BOTH, expand=True, pady=15)
        self._aplicar_cantos_arredondados(self.frame_lista, 10)
    
        # Listbox com scrollbar e visual melhorado
        self.listbox = tk.Listbox(
            self.frame_lista, 
            font=('Roboto', 10),
            bg='white',
            fg=self.cores['texto'],
            borderwidth=0,
            highlightthickness=0,
            selectbackground=self.cores['azul'],
            selectforeground='white'
        )
        scrollbar = Scrollbar(self.frame_lista)
        self.listbox.config(yscrollcommand=scrollbar.set)
        scrollbar.config(command=self.listbox.yview)
        
        self.listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
    
        # Rodapé com estilo Chromium
        frame_rodape = ttk.Frame(self.root, style='Footer.TFrame')
        frame_rodape.pack(side=tk.BOTTOM, fill=tk.X)
    
        btn_sair = ttk.Button(
            frame_rodape, 
            text="Sair", 
            command=self.root.quit,
            style='Red.TButton'
        )
        self._aplicar_cantos_arredondados(btn_sair, 20)
        btn_sair.pack(pady=15)
    
    def _aplicar_cantos_arredondados(self, widget, raio=10):
        """Aplica cantos arredondados a um widget."""
        try:
            # Tenta aplicar cantos arredondados ao widget
            # Nem todos os widgets suportam essa funcionalidade
            widget.configure(style="Rounded")
        except:
            pass
        
    def atualizar_lista_cacambas(self) -> None:
        """Atualiza a lista de caçambas na interface."""
        self.listbox.delete(0, tk.END)
        
        cacambas = GerenciadorArquivos.carregar_cacambas()
        for cacamba in cacambas:
            # Formata a exibição na lista
            status = ""
            if cacamba.precisa_retirada:
                status = f" [RETIRAR - {cacamba.dias_no_local} dias]"
                
            texto = f"Caçamba {cacamba.numero} - {cacamba.endereco_completo}{status}"
            self.listbox.insert(tk.END, texto)
    
    def verificar_e_notificar_retiradas(self) -> None:
        """Verifica e notifica sobre caçambas prontas para retirada."""
        cacambas_para_retirada = self.gerenciador.verificar_cacambas_para_retirada()
        
        # Exibe notificações para caçambas que precisam ser retiradas
        for cacamba in cacambas_para_retirada:
            messagebox.showwarning(
                "ALERTA", 
                f"A caçamba {cacamba.numero} localizada em {cacamba.endereco_completo} " +
                f"está no local há {cacamba.dias_no_local} dias e está pronta para retirada. " +
                f"Data de colocação: {cacamba.data_colocacao}"
            )
    
    def gerar_e_mostrar_mapa(self) -> None:
        """Gera e abre o mapa com as localizações das caçambas."""
        cacambas = GerenciadorArquivos.carregar_cacambas()
        arquivo_mapa = ServicoLocalizacao.gerar_mapa(cacambas)
        
        if arquivo_mapa:
            # Abre o arquivo do mapa no navegador padrão
            try:
                import webbrowser
                caminho_completo = os.path.abspath(arquivo_mapa)
                webbrowser.open(f'file://{caminho_completo}')
            except Exception as e:
                print(Fore.RED + f"Erro ao abrir o mapa: {e}")
                messagebox.showinfo(
                    "Informação", 
                    f"Mapa salvo em {arquivo_mapa}. Abra-o manualmente em seu navegador."
                )
    
    def iniciar(self) -> None:
        """Inicia a execução da interface gráfica."""
        # Verifica caçambas para retirada na inicialização
        self.verificar_e_notificar_retiradas()
        
        # Inicia o loop principal
        self.root.mainloop()


def main():
    """Função principal que inicia a aplicação."""
    # Verifica/cria arquivo necessário
    GerenciadorArquivos.criar_arquivo_se_nao_existir()
    
    # Inicializa o gerenciador
    gerenciador = GerenciadorCacambas()
    
    # Inicializa a interface
    interface = InterfaceGrafica(gerenciador)
    
    # Inicia a aplicação
    interface.iniciar()


if __name__ == '__main__':
    main()